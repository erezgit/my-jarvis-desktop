"""
FastAPI Excel Processing Backend
Hybrid architecture using pandas + openpyxl for optimal performance
"""

from fastapi import FastAPI, File, UploadFile, HTTPException, WebSocket, WebSocketDisconnect
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
import asyncio
import json
import os
import uuid
from pathlib import Path
from typing import Dict, List, Optional
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from watchfiles import awatch
import logging

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = FastAPI(title="Excel Editor API", version="1.0.0")

# CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Configure properly for production
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configuration
UPLOAD_DIR = Path("./uploads")
UPLOAD_DIR.mkdir(exist_ok=True)

# WebSocket connection manager
class ConnectionManager:
    def __init__(self):
        self.active_connections: Dict[str, WebSocket] = {}

    async def connect(self, websocket: WebSocket, session_id: str):
        await websocket.accept()
        self.active_connections[session_id] = websocket
        logger.info(f"WebSocket connected: {session_id}")

    def disconnect(self, session_id: str):
        if session_id in self.active_connections:
            del self.active_connections[session_id]
            logger.info(f"WebSocket disconnected: {session_id}")

    async def send_message(self, session_id: str, message: dict):
        if session_id in self.active_connections:
            try:
                await self.active_connections[session_id].send_text(json.dumps(message))
            except Exception as e:
                logger.error(f"Error sending message to {session_id}: {e}")
                self.disconnect(session_id)

manager = ConnectionManager()

class ExcelProcessor:
    """Hybrid Excel processor using pandas for analysis and openpyxl for Excel I/O"""

    def __init__(self, file_path: Path):
        self.file_path = file_path
        self.workbook = None
        self.data_cache = {}

    def load_workbook(self):
        """Load workbook with formula preservation"""
        try:
            self.workbook = openpyxl.load_workbook(self.file_path, data_only=False)
            logger.info(f"Loaded workbook with {len(self.workbook.sheetnames)} sheets")
            return True
        except Exception as e:
            logger.error(f"Error loading workbook: {e}")
            return False

    def get_sheet_data(self, sheet_name: str = None) -> dict:
        """Get sheet data optimized for frontend rendering"""
        if not self.workbook:
            return {"error": "Workbook not loaded"}

        sheet_name = sheet_name or self.workbook.active.title
        worksheet = self.workbook[sheet_name]

        # Use pandas for efficient data extraction
        data = []
        for row in worksheet.iter_rows(values_only=False):
            row_data = []
            for cell in row:
                cell_info = {
                    "value": cell.value,
                    "formula": cell.formula if cell.formula else None,
                    "data_type": cell.data_type,
                    "coordinate": cell.coordinate
                }
                # Add formatting information if available
                if cell.font:
                    cell_info["font"] = {
                        "name": cell.font.name,
                        "size": cell.font.size,
                        "bold": cell.font.bold,
                        "italic": cell.font.italic
                    }
                if cell.fill and cell.fill.start_color:
                    cell_info["fill_color"] = cell.fill.start_color.rgb

                row_data.append(cell_info)
            data.append(row_data)

        return {
            "sheet_name": sheet_name,
            "data": data,
            "max_row": worksheet.max_row,
            "max_column": worksheet.max_column,
            "sheet_names": self.workbook.sheetnames
        }

    def update_cell(self, sheet_name: str, row: int, col: int, value: str, formula: str = None):
        """Update cell value while preserving formulas"""
        if not self.workbook:
            return False

        try:
            worksheet = self.workbook[sheet_name]
            cell = worksheet.cell(row=row, column=col)

            if formula:
                cell.value = formula
            else:
                cell.value = value

            return True
        except Exception as e:
            logger.error(f"Error updating cell: {e}")
            return False

    def save_workbook(self) -> bool:
        """Save workbook with formula preservation"""
        try:
            self.workbook.save(self.file_path)
            logger.info(f"Saved workbook to {self.file_path}")
            return True
        except Exception as e:
            logger.error(f"Error saving workbook: {e}")
            return False

    def analyze_with_pandas(self, sheet_name: str = None) -> dict:
        """Use pandas for data analysis while preserving Excel structure"""
        try:
            # Read with pandas for analysis
            df = pd.read_excel(self.file_path, sheet_name=sheet_name, engine='openpyxl')

            analysis = {
                "shape": df.shape,
                "columns": df.columns.tolist(),
                "dtypes": df.dtypes.to_dict(),
                "null_counts": df.isnull().sum().to_dict(),
                "summary_stats": df.describe().to_dict() if df.select_dtypes(include='number').shape[1] > 0 else {}
            }

            return analysis
        except Exception as e:
            logger.error(f"Error in pandas analysis: {e}")
            return {"error": str(e)}

# File watching system
class FileWatcher:
    def __init__(self, file_path: Path, session_id: str):
        self.file_path = file_path
        self.session_id = session_id
        self.watching = False

    async def start_watching(self):
        """Start watching file for changes"""
        self.watching = True
        try:
            async for changes in awatch(self.file_path.parent):
                if not self.watching:
                    break

                for change_type, changed_path in changes:
                    if Path(changed_path) == self.file_path:
                        logger.info(f"File changed: {changed_path}")
                        await manager.send_message(self.session_id, {
                            "type": "file_changed",
                            "path": str(changed_path),
                            "change_type": change_type.name
                        })
        except Exception as e:
            logger.error(f"Error in file watcher: {e}")

    def stop_watching(self):
        self.watching = False

# Global file watchers
file_watchers: Dict[str, FileWatcher] = {}

@app.post("/upload")
async def upload_excel_file(file: UploadFile = File(...)):
    """Upload and validate Excel file"""
    if not file.filename.endswith(('.xlsx', '.xlsm', '.xls')):
        raise HTTPException(status_code=400, detail="Invalid file format")

    # Generate unique session ID
    session_id = str(uuid.uuid4())
    file_path = UPLOAD_DIR / f"{session_id}_{file.filename}"

    # Security: Validate file size
    content = await file.read()
    if len(content) > 50 * 1024 * 1024:  # 50MB limit
        raise HTTPException(status_code=400, detail="File too large")

    # Save file
    with open(file_path, "wb") as f:
        f.write(content)

    # Validate it's a proper Excel file
    processor = ExcelProcessor(file_path)
    if not processor.load_workbook():
        os.remove(file_path)
        raise HTTPException(status_code=400, detail="Invalid Excel file")

    return {
        "session_id": session_id,
        "filename": file.filename,
        "file_path": str(file_path),
        "sheets": processor.workbook.sheetnames
    }

@app.get("/sheet/{session_id}")
async def get_sheet_data(session_id: str, sheet_name: str = None):
    """Get sheet data for preview"""
    file_path = None
    for file in UPLOAD_DIR.glob(f"{session_id}_*"):
        file_path = file
        break

    if not file_path or not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found")

    processor = ExcelProcessor(file_path)
    if not processor.load_workbook():
        raise HTTPException(status_code=500, detail="Error loading file")

    return processor.get_sheet_data(sheet_name)

@app.post("/update/{session_id}")
async def update_cell_data(session_id: str, update_data: dict):
    """Update cell data and save"""
    file_path = None
    for file in UPLOAD_DIR.glob(f"{session_id}_*"):
        file_path = file
        break

    if not file_path or not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found")

    processor = ExcelProcessor(file_path)
    if not processor.load_workbook():
        raise HTTPException(status_code=500, detail="Error loading file")

    # Update cell
    success = processor.update_cell(
        update_data["sheet_name"],
        update_data["row"],
        update_data["col"],
        update_data["value"],
        update_data.get("formula")
    )

    if success:
        processor.save_workbook()
        return {"success": True}
    else:
        raise HTTPException(status_code=500, detail="Error updating cell")

@app.get("/analyze/{session_id}")
async def analyze_sheet(session_id: str, sheet_name: str = None):
    """Analyze sheet data with pandas"""
    file_path = None
    for file in UPLOAD_DIR.glob(f"{session_id}_*"):
        file_path = file
        break

    if not file_path or not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found")

    processor = ExcelProcessor(file_path)
    return processor.analyze_with_pandas(sheet_name)

@app.get("/download/{session_id}")
async def download_file(session_id: str):
    """Download modified Excel file"""
    file_path = None
    for file in UPLOAD_DIR.glob(f"{session_id}_*"):
        file_path = file
        break

    if not file_path or not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found")

    return FileResponse(
        path=file_path,
        filename=file_path.name.split("_", 1)[1],  # Remove session ID from filename
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.websocket("/ws/{session_id}")
async def websocket_endpoint(websocket: WebSocket, session_id: str):
    """WebSocket for real-time updates"""
    await manager.connect(websocket, session_id)

    # Start file watcher
    file_path = None
    for file in UPLOAD_DIR.glob(f"{session_id}_*"):
        file_path = file
        break

    if file_path:
        watcher = FileWatcher(file_path, session_id)
        file_watchers[session_id] = watcher

        # Start watching in background
        watch_task = asyncio.create_task(watcher.start_watching())

    try:
        while True:
            data = await websocket.receive_text()
            message = json.loads(data)

            # Handle different message types
            if message.get("type") == "ping":
                await manager.send_message(session_id, {"type": "pong"})

    except WebSocketDisconnect:
        manager.disconnect(session_id)
        if session_id in file_watchers:
            file_watchers[session_id].stop_watching()
            del file_watchers[session_id]
        if 'watch_task' in locals():
            watch_task.cancel()

@app.on_event("startup")
async def startup_event():
    logger.info("Excel Editor API started")

@app.on_event("shutdown")
async def shutdown_event():
    # Stop all file watchers
    for watcher in file_watchers.values():
        watcher.stop_watching()
    logger.info("Excel Editor API shutdown")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)